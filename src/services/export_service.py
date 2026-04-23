"""
Export service — формати: CSV, XLSX, TXT (pipe-delimited).

Відповідає за:
  - Перетворення ORM-моделей у rows-of-dicts
  - Рендеринг у потрібний формат
  - Повернення (content_bytes, mime_type, file_extension) для response
"""
from __future__ import annotations

import csv
import io
from datetime import datetime
from typing import Any, Iterable, List, Literal, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from db.models import Call, User

ExportFormat = Literal["csv", "xlsx", "txt"]

# ── Column definitions ────────────────────────────────────────────────────────

CALL_COLUMNS: Sequence[Tuple[str, str]] = [
    ("id",               "ID"),
    ("call_type",        "Type"),
    ("call_state",       "State"),
    ("from_number",      "From"),
    ("to_number",        "To"),
    ("date",             "Date"),
    ("seconds_fulltime", "Full Time (s)"),
    ("seconds_talktime", "Talk Time (s)"),
    ("callback",         "Callback"),
    ("mp3_link",         "Recording URL"),
    ("user_phone",       "Client Phone"),
    ("user_name",        "Client Name"),
]

USER_COLUMNS: Sequence[Tuple[str, str]] = [
    ("id",            "ID"),
    ("phone_number",  "Phone"),
    ("name",          "Name"),
    ("category",      "Category"),
    ("types",         "Types"),
    ("calls_count",   "Calls"),
    ("description",   "Notes"),
    ("created_at",    "Created At"),
]


# ── Row transformers ──────────────────────────────────────────────────────────

def _fmt_dt(v: datetime | None) -> str:
    return v.strftime("%Y-%m-%d %H:%M:%S") if v else ""


def _call_to_row(c: Call) -> dict:
    # enum values: CallType.IN → "IN" (take .value)
    ctype = getattr(c.call_type, "value", c.call_type) if c.call_type else ""
    cstate = getattr(c.call_state, "value", c.call_state) if c.call_state else ""
    return {
        "id":               c.id,
        "call_type":        ctype or "",
        "call_state":       cstate or "",
        "from_number":      c.from_number or "",
        "to_number":        c.to_number or "",
        "date":             _fmt_dt(c.date),
        "seconds_fulltime": round(c.seconds_fulltime or 0, 1),
        "seconds_talktime": round(c.seconds_talktime or 0, 1),
        "callback":         "Yes" if c.callback else "No",
        "mp3_link":         c.mp3_link or "",
        "user_phone":       (c.user.phone_number if c.user else "") or "",
        "user_name":        (c.user.name if c.user else "") or "",
    }


def _user_to_row(u: User) -> dict:
    return {
        "id":           u.id,
        "phone_number": u.phone_number or "",
        "name":         u.name or "",
        "category":     u.category.name if u.category else "",
        "types":        ", ".join(t.name for t in (u.types or [])),
        "calls_count":  u.calls_count or 0,
        "description":  (u.description or "").replace("\n", " ").strip(),
        "created_at":   _fmt_dt(u.created_at),
    }


# ── Renderers ─────────────────────────────────────────────────────────────────

def _render_csv(headers: Sequence[str], keys: Sequence[str], rows: Iterable[dict]) -> bytes:
    # BOM для коректного відкриття у Excel
    buf = io.StringIO()
    buf.write("﻿")
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(k, "") for k in keys])
    return buf.getvalue().encode("utf-8")


def _render_txt(headers: Sequence[str], keys: Sequence[str], rows: List[dict]) -> bytes:
    """Pipe-delimited з вирівнюванням колонок за шириною."""
    # обчислюємо ширину для вирівнювання
    widths = [len(h) for h in headers]
    for r in rows:
        for i, k in enumerate(keys):
            widths[i] = max(widths[i], len(str(r.get(k, ""))))

    def fmt(values: Sequence[Any]) -> str:
        return "| " + " | ".join(
            str(v).ljust(widths[i]) for i, v in enumerate(values)
        ) + " |"

    separator = "+" + "+".join("-" * (w + 2) for w in widths) + "+"
    lines: List[str] = [separator, fmt(headers), separator]
    for r in rows:
        lines.append(fmt([r.get(k, "") for k in keys]))
    lines.append(separator)
    return ("\n".join(lines) + "\n").encode("utf-8")


def _render_xlsx(headers: Sequence[str], keys: Sequence[str], rows: Iterable[dict], title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # обмеження Excel на довжину назви sheet

    # Header row — жирний, темно-синій фон, білий текст
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center = Alignment(horizontal="center", vertical="center")

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center

    # Data rows
    for row_idx, r in enumerate(rows, start=2):
        for col_idx, k in enumerate(keys, start=1):
            ws.cell(row=row_idx, column=col_idx, value=r.get(k, ""))

    # Auto-width (приблизно) + freeze header
    for i, h in enumerate(headers, 1):
        max_len = len(h)
        col_letter = get_column_letter(i)
        for cell in ws[col_letter]:
            val = str(cell.value) if cell.value is not None else ""
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = "A2"

    # Filter (autofilter) на весь діапазон
    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Public API ────────────────────────────────────────────────────────────────

def _mime_for(fmt: ExportFormat) -> Tuple[str, str]:
    """Returns (mime_type, file_extension)."""
    return {
        "csv":  ("text/csv; charset=utf-8",                     "csv"),
        "xlsx": ("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"),
        "txt":  ("text/plain; charset=utf-8",                   "txt"),
    }[fmt]


def export_calls(calls: List[Call], fmt: ExportFormat) -> Tuple[bytes, str, str]:
    keys = [k for k, _ in CALL_COLUMNS]
    headers = [h for _, h in CALL_COLUMNS]
    rows = [_call_to_row(c) for c in calls]

    if fmt == "csv":
        content = _render_csv(headers, keys, rows)
    elif fmt == "txt":
        content = _render_txt(headers, keys, rows)
    else:  # xlsx
        content = _render_xlsx(headers, keys, rows, title="Calls")

    mime, ext = _mime_for(fmt)
    filename = f"calls_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
    return content, mime, filename


def export_users(users: List[User], fmt: ExportFormat) -> Tuple[bytes, str, str]:
    keys = [k for k, _ in USER_COLUMNS]
    headers = [h for _, h in USER_COLUMNS]
    rows = [_user_to_row(u) for u in users]

    if fmt == "csv":
        content = _render_csv(headers, keys, rows)
    elif fmt == "txt":
        content = _render_txt(headers, keys, rows)
    else:  # xlsx
        content = _render_xlsx(headers, keys, rows, title="Clients")

    mime, ext = _mime_for(fmt)
    filename = f"clients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{ext}"
    return content, mime, filename
